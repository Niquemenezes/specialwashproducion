import io
from datetime import datetime

from flask import Blueprint, send_file
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import func

from extensions import db
from models import Cliente, Coche, GastoEmpresa, InspeccionRecepcion, Servicio
from utils.auth_utils import role_required

export_bp = Blueprint("export", __name__)

# ── Colores SpecialWash ──────────────────────────────────────────────────────
C_DARK   = "1B2A4A"
C_GOLD   = "D4AF37"
C_WHITE  = "FFFFFF"
C_GRAY   = "F8FAFC"
C_LIGHT  = "EFF6FF"
C_GREEN  = "D1FAE5"
C_RED    = "FEE2E2"
C_ORANGE = "FFEDD5"
C_BORDER = "E2E8F0"


def _fill(color):
    return PatternFill("solid", start_color=color, fgColor=color)


def _thin():
    s = Side(style="thin", color=C_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)


def _header_cell(ws, coord, value, bg=C_DARK, fg=C_WHITE, size=10, bold=True):
    c = ws[coord]
    c.value = value
    c.font = Font(name="Arial", bold=bold, size=size, color=fg)
    c.fill = _fill(bg)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = _thin()
    return c


def _data_cell(ws, coord, value, bg=C_WHITE, fg="1B2A4A", size=9,
               bold=False, h_align="center", number_format=None):
    c = ws[coord]
    c.value = value
    c.font = Font(name="Arial", bold=bold, size=size, color=fg)
    c.fill = _fill(bg)
    c.alignment = Alignment(horizontal=h_align, vertical="center")
    c.border = _thin()
    if number_format:
        c.number_format = number_format
    return c


def _mes_label(mes):
    nombres = ["", "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
               "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
    return nombres[mes] if 1 <= mes <= 12 else str(mes)


# ── Hoja: portada / dashboard ────────────────────────────────────────────────
def _build_dashboard_sheet(wb, anio, servicios_rows, gastos_total):
    ws = wb.create_sheet("Dashboard")

    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 16
    ws.column_dimensions["G"].width = 16

    # Título
    ws.row_dimensions[2].height = 36
    ws.merge_cells("B2:G2")
    ws["B2"].value = f"SPECIAL WASH AUTO SPA — RESUMEN {anio}"
    ws["B2"].font = Font(name="Arial", bold=True, size=14, color=C_WHITE)
    ws["B2"].fill = _fill(C_DARK)
    ws["B2"].alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[3].height = 14
    ws.merge_cells("B3:G3")
    ws["B3"].value = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}  ·  specialwash.studio"
    ws["B3"].font = Font(name="Arial", size=9, color="94A3B8", italic=True)
    ws["B3"].fill = _fill(C_DARK)
    ws["B3"].alignment = Alignment(horizontal="center", vertical="center")

    # KPI cards
    total_anio = sum(r.total or 0 for r in servicios_rows)
    trabajos_anio = sum(r.count or 0 for r in servicios_rows)
    beneficio = total_anio - gastos_total

    kpis = [
        ("FACTURADO AÑO", f"{total_anio:,.2f} €", f"{trabajos_anio} trabajos"),
        ("GASTOS AÑO",    f"{gastos_total:,.2f} €", "registrados"),
        ("BENEFICIO EST.", f"{beneficio:,.2f} €",
         "▲ positivo" if beneficio >= 0 else "▼ negativo"),
    ]
    kpi_cols = [("B", "C"), ("D", "E"), ("F", "G")]

    for i, (title, val, sub) in enumerate(kpis):
        sc, ec = kpi_cols[i]
        for r in range(5, 9):
            ws.row_dimensions[r].height = 18 if r == 6 else 14
            ws.merge_cells(f"{sc}{r}:{ec}{r}")
        ws[f"{sc}5"].value = title
        ws[f"{sc}5"].font = Font(name="Arial", bold=True, size=8, color="64748B")
        ws[f"{sc}5"].fill = _fill(C_GRAY)
        ws[f"{sc}5"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"{sc}6"].value = val
        ws[f"{sc}6"].font = Font(name="Arial", bold=True, size=14, color=C_DARK)
        ws[f"{sc}6"].fill = _fill(C_WHITE)
        ws[f"{sc}6"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"{sc}7"].value = sub
        ws[f"{sc}7"].font = Font(name="Arial", size=8, color="94A3B8")
        ws[f"{sc}7"].fill = _fill(C_WHITE)
        ws[f"{sc}7"].alignment = Alignment(horizontal="center", vertical="center")
        for r in range(5, 9):
            for col in [sc, ec]:
                ws[f"{col}{r}"].border = _thin()

    # Tabla mensual
    ws.row_dimensions[10].height = 18
    ws.merge_cells("B10:F10")
    ws["B10"].value = "FACTURACIÓN MENSUAL"
    ws["B10"].font = Font(name="Arial", bold=True, size=10, color=C_WHITE)
    ws["B10"].fill = _fill(C_DARK)
    ws["B10"].alignment = Alignment(horizontal="left", vertical="center", indent=1)

    headers = ["Mes", "Trabajos", "Facturado (€)", "IVA 21% (€)", "Base (€)"]
    h_cols = ["B", "C", "D", "E", "F"]
    ws.row_dimensions[11].height = 16
    for h, col in zip(headers, h_cols):
        _header_cell(ws, f"{col}11", h, bg=C_GRAY, fg="64748B", bold=True)

    totales_mes = {int(r.mes): {"total": float(r.total or 0), "count": int(r.count)} for r in servicios_rows}
    row_colors = [C_WHITE, C_GRAY]
    grand_total = grand_count = grand_iva = 0

    for m in range(1, 13):
        r = 12 + m - 1
        ws.row_dimensions[r].height = 15
        data = totales_mes.get(m, {"total": 0, "count": 0})
        total = data["total"]
        iva = round(total * 0.21, 2)
        base = round(total - iva, 2)
        grand_total += total
        grand_count += data["count"]
        grand_iva += iva
        bg = row_colors[m % 2]
        vals = [_mes_label(m), data["count"], total, iva, base]
        for v, col in zip(vals, h_cols):
            fmt = "#,##0.00 €" if col in ("D", "E", "F") else None
            _data_cell(ws, f"{col}{r}", v, bg=bg, number_format=fmt)

    # Total row
    r_tot = 24
    ws.row_dimensions[r_tot].height = 18
    for v, col in zip(["TOTAL", grand_count, grand_total, grand_iva,
                        round(grand_total - grand_iva, 2)], h_cols):
        fmt = "#,##0.00 €" if col in ("D", "E", "F") else None
        _data_cell(ws, f"{col}{r_tot}", v, bg=C_DARK, fg=C_WHITE,
                   bold=True, number_format=fmt)

    ws.freeze_panes = "B5"
    return ws


# ── Hoja: mes individual ─────────────────────────────────────────────────────
def _build_mes_sheet(wb, mes_label, servicios):
    ws = wb.create_sheet(mes_label)

    col_widths = {"A": 12, "B": 18, "C": 22, "D": 13,
                  "E": 38, "F": 11, "G": 9, "H": 14, "I": 13, "J": 22}
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

    # Título
    ws.row_dimensions[1].height = 22
    ws.merge_cells("A1:J1")
    ws["A1"].value = mes_label
    ws["A1"].font = Font(name="Arial", bold=True, size=13, color=C_DARK)
    ws["A1"].alignment = Alignment(vertical="center", indent=1)

    # Headers
    headers = ["Fecha", "Modelo/Marca", "Cliente", "Matrícula",
               "Tipo de Servicio", "Precio", "IVA", "Método Pago", "Entrega", "Observaciones"]
    ws.row_dimensions[2].height = 18
    for i, h in enumerate(headers):
        col = get_column_letter(i + 1)
        _header_cell(ws, f"{col}2", h, bg=C_DARK, fg=C_GOLD)

    # Datos
    row_colors = [C_WHITE, C_GRAY]
    for i, svc in enumerate(servicios):
        r = 3 + i
        ws.row_dimensions[r].height = 15
        bg = row_colors[i % 2]
        fecha = svc.fecha.strftime("%d/%m/%Y") if svc.fecha else ""
        modelo = f"{svc.coche.marca or ''} {svc.coche.modelo or ''}".strip() if svc.coche else ""
        cliente = svc.coche.cliente.nombre if svc.coche and svc.coche.cliente else ""
        matricula = svc.coche.matricula if svc.coche else ""
        iva = round(float(svc.precio or 0) * 0.21, 2)

        row_vals = [fecha, modelo, cliente, matricula,
                    svc.tipo_servicio, float(svc.precio or 0), iva,
                    "", "", svc.observaciones or ""]
        for j, v in enumerate(row_vals):
            col = get_column_letter(j + 1)
            fmt = "#,##0.00" if j in (5, 6) else None
            align = "left" if j in (1, 2, 4, 9) else "center"
            _data_cell(ws, f"{col}{r}", v, bg=bg, h_align=align, number_format=fmt)

    # Total
    if servicios:
        r_tot = 3 + len(servicios)
        ws.row_dimensions[r_tot].height = 18
        total = sum(float(s.precio or 0) for s in servicios)
        ws.merge_cells(f"A{r_tot}:E{r_tot}")
        _data_cell(ws, f"A{r_tot}", "TOTAL MES", bg=C_DARK, fg=C_WHITE, bold=True)
        _data_cell(ws, f"F{r_tot}", total, bg=C_DARK, fg=C_GOLD,
                   bold=True, number_format="#,##0.00 €")
        for col in ["G", "H", "I", "J"]:
            _data_cell(ws, f"{col}{r_tot}", "", bg=C_DARK)

    ws.freeze_panes = "A3"
    return ws


# ── Hoja: top clientes ───────────────────────────────────────────────────────
def _build_clientes_sheet(wb, anio):
    ws = wb.create_sheet("Top Clientes")

    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 3

    ws.row_dimensions[2].height = 22
    ws.merge_cells("B2:E2")
    ws["B2"].value = f"TOP CLIENTES — {anio}"
    ws["B2"].font = Font(name="Arial", bold=True, size=11, color=C_WHITE)
    ws["B2"].fill = _fill(C_DARK)
    ws["B2"].alignment = Alignment(horizontal="center", vertical="center")

    for h, col in zip(["Cliente", "Trabajos", "Total (€)", "Ticket Medio (€)"],
                      ["B", "C", "D", "E"]):
        _header_cell(ws, f"{col}3", h, bg="1E3A5F", fg=C_WHITE)

    rows = (
        db.session.query(
            Cliente.nombre,
            func.count(Servicio.id).label("count"),
            func.sum(Servicio.precio).label("total"),
        )
        .join(Coche, Coche.cliente_id == Cliente.id)
        .join(Servicio, Servicio.coche_id == Coche.id)
        .filter(func.strftime("%Y", Servicio.fecha) == str(anio))
        .group_by(Cliente.id, Cliente.nombre)
        .order_by(func.sum(Servicio.precio).desc())
        .limit(20)
        .all()
    )

    for i, row in enumerate(rows):
        r = 4 + i
        ws.row_dimensions[r].height = 15
        bg = C_WHITE if i % 2 == 0 else C_GRAY
        total = float(row.total or 0)
        ticket = round(total / int(row.count), 2) if row.count else 0
        for v, col, fmt in zip(
            [row.nombre, int(row.count), total, ticket],
            ["B", "C", "D", "E"],
            [None, None, "#,##0.00 €", "#,##0.00 €"],
        ):
            align = "left" if col == "B" else "center"
            _data_cell(ws, f"{col}{r}", v, bg=bg, h_align=align, number_format=fmt)

    return ws


# ── Ruta principal ───────────────────────────────────────────────────────────
@export_bp.route("/api/export/excel", methods=["GET"])
@role_required("administrador")
def export_excel():
    """Genera y descarga el Excel anual de SpecialWash."""
    from flask import request as req
    anio = req.args.get("anio", datetime.now().year, type=int)

    wb = Workbook()
    # Eliminar hoja por defecto
    wb.remove(wb.active)

    # Datos anuales de servicios agrupados por mes
    servicios_rows = (
        db.session.query(
            func.strftime("%m", Servicio.fecha).label("mes"),
            func.sum(Servicio.precio).label("total"),
            func.count(Servicio.id).label("count"),
        )
        .filter(func.strftime("%Y", Servicio.fecha) == str(anio))
        .group_by(func.strftime("%m", Servicio.fecha))
        .all()
    )

    gastos_total = float(
        db.session.query(func.sum(GastoEmpresa.importe))
        .filter(func.strftime("%Y", GastoEmpresa.fecha) == str(anio))
        .scalar() or 0
    )

    # 1. Dashboard
    _build_dashboard_sheet(wb, anio, servicios_rows, gastos_total)

    # 2. Hoja por cada mes con datos
    for mes in range(1, 13):
        servicios = (
            Servicio.query
            .filter(
                func.strftime("%Y", Servicio.fecha) == str(anio),
                func.strftime("%m", Servicio.fecha) == f"{mes:02d}",
            )
            .order_by(Servicio.fecha.asc())
            .all()
        )
        _build_mes_sheet(wb, _mes_label(mes), servicios)

    # 3. Top clientes
    _build_clientes_sheet(wb, anio)

    # Colores de pestañas
    tab_colors = {
        "Dashboard": "1B2A4A",
        "Enero": "1D4ED8", "Febrero": "0891B2", "Marzo": "059669",
        "Abril": "65A30D", "Mayo": "CA8A04", "Junio": "EA580C",
        "Julio": "DC2626", "Agosto": "DB2777", "Septiembre": "9333EA",
        "Octubre": "7C3AED", "Noviembre": "2563EB", "Diciembre": "0F172A",
        "Top Clientes": "D4AF37",
    }
    for name, color in tab_colors.items():
        if name in wb.sheetnames:
            wb[name].sheet_properties.tabColor = color

    # Guardar en memoria y enviar
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"SpecialWash_{anio}_{datetime.now().strftime('%Y%m%d')}.xlsx"

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )